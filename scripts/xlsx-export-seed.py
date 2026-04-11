"""Export matex_seed_data.xlsx sheets to JSON on stdout (requires openpyxl)."""
import json
import sys

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "matex_seed_data.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        out[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    print(json.dumps(out, default=str))


if __name__ == "__main__":
    main()
